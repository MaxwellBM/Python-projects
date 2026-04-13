"""
excel_processor.py — Procesa los datos extraídos por RPA y genera un reporte Excel
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import logging

from rpa import TableRow

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path("data")


class ExcelReporter:
    """Genera un reporte Excel formateado a partir de los datos RPA."""

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    THIN = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def __init__(self, output_path: Path = OUTPUT_DIR / "rpa_report.xlsx"):
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def _style_header(self, ws, row: int, cols: int) -> None:
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER

    def _auto_width(self, ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    def generate(self, table_rows: list[TableRow], rpa_results: dict) -> Path:
        wb = Workbook()

        # ── Hoja 1: Datos de tabla ─────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Datos Extraídos"

        headers = ["Apellido", "Nombre", "Email", "Deuda", "Sitio Web", "Acción"]
        for col, h in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=h)
        self._style_header(ws1, 1, len(headers))

        for i, row in enumerate(table_rows, 2):
            values = [row.last_name, row.first_name, row.email,
                      row.due, row.web_site, row.action]
            for col, val in enumerate(values, 1):
                cell = ws1.cell(row=i, column=col, value=val)
                cell.border = self.BORDER
                if i % 2 == 0:
                    cell.fill = self.ALT_FILL

        self._auto_width(ws1)

        # ── Hoja 2: Resumen de ejecución ───────────────────────────────────
        ws2 = wb.create_sheet("Resumen Ejecución")
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 45

        title_font = Font(bold=True, size=14, color="1F4E79")
        ws2["A1"] = "Reporte de Automatización RPA"
        ws2["A1"].font = title_font
        ws2.merge_cells("A1:B1")
        ws2["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws2["A2"].font = Font(italic=True, color="888888")
        ws2.merge_cells("A2:B2")

        data = [
            ("", ""),
            ("MÉTRICAS", ""),
            ("Filas extraídas", len(table_rows)),
            ("Login exitoso", "✅" if rpa_results.get("login", {}).get("success") else "❌"),
            ("Alerts manejados", len(rpa_results.get("alerts", {}))),
            ("Reporte descargado", "✅" if rpa_results.get("report", {}).get("saved") else "❌"),
            ("", ""),
            ("DETALLES LOGIN", ""),
            ("Mensaje", rpa_results.get("login", {}).get("message", "N/A")),
            ("", ""),
            ("ALERTS JS", ""),
        ]

        for alert_key, alert_val in rpa_results.get("alerts", {}).items():
            data.append((alert_key, str(alert_val)))

        for row_idx, (key, val) in enumerate(data, 3):
            ws2.cell(row=row_idx, column=1, value=key)
            ws2.cell(row=row_idx, column=2, value=val)
            if key in ("MÉTRICAS", "DETALLES LOGIN", "ALERTS JS"):
                ws2.cell(row=row_idx, column=1).font = Font(bold=True, color="1F4E79")

        logger.info(f"Reporte Excel generado: {self.output_path}")
        wb.save(self.output_path)
        return self.output_path


if __name__ == "__main__":
    from rpa import RPABot
    with RPABot(headless=True) as bot:
        results = bot.run_full_pipeline()
    rows = [
        TableRow("Smith", "John", "jsmith@example.com", "$50.00", "http://example.com", "edit delete"),
    ]
    reporter = ExcelReporter()
    path = reporter.generate(rows, results)
    print(f"Excel guardado en: {path}")
